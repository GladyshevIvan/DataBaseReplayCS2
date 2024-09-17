import os
from pandas import DataFrame, ExcelWriter
from openpyxl import load_workbook
from ParseDemo import StatsFromDemo
from ConnectToDataBase import CS2DataBase


class ActionWithData:
    def __init__(self, raw_dem, date_and_time, database_action_choice, to_stream_choice):
        self.raw_dem = raw_dem
        self.date_and_time = date_and_time
        self.to_stream_choice = to_stream_choice
        match database_action_choice:
            case 'Ничего не делать':
                self.get_stats_from_replay()
            case 'Извлечь матч':
                self.get_stats_from_database()
            case 'Записать матч':
                self.insert_stats_to_database()

    #Получение данных из повтора
    def get_stats_from_replay(self):
        cs2_for_excel = StatsFromDemo(self.raw_dem, self.date_and_time)
        self.players_stats_df = DataFrame.from_dict(cs2_for_excel.players_stats, orient='index').sort_values(by=['team', 'kills'], ascending=False)
        self.total_score_df = DataFrame.from_dict(cs2_for_excel.total_score, orient='index').sort_values(by='team', ascending=False)
        self.make_datasource_excel()

    #Получение данных из базы данных
    def get_stats_from_database(self):
        cs2_for_excel = CS2DataBase()
        a = cs2_for_excel.extract_one_game_from_database(self.raw_dem, self.date_and_time)
        self.players_stats_df = DataFrame.from_dict(a[1], orient='index').sort_values(by=['team', 'kills'], ascending=False)
        self.total_score_df = DataFrame.from_dict(a[0], orient='index').sort_values(by='team', ascending=False)
        self.make_datasource_excel()

    def insert_stats_to_database(self):
        cs2_for_excel = CS2DataBase()
        cs2_for_excel.insert_cs2_stats_to_db(self.raw_dem, self.date_and_time)
        if self.to_stream_choice:
            self.get_stats_from_database()


    def make_datasource_excel(self):
        if os.path.isfile('CS2 stats for stream.xlsx'):
            from openpyxl import load_workbook
            file = load_workbook('CS2 stats for stream.xlsx')

            if 'CS2' in file.sheetnames:
                sheet = file['CS2']
            else:
                sheet = file.create_sheet(title='CS2')

            with ExcelWriter('CS2 stats for stream.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                self.players_stats_df.to_excel(
                    writer,
                    sheet_name='CS2',
                    index=False,
                    startrow=0,
                    startcol=0,
                )

            with ExcelWriter('CS2 stats for stream.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                self.total_score_df.to_excel(
                    writer,
                    sheet_name='CS2',
                    index=False,
                    startrow=0,
                    startcol=8,
                )

        else:
            with ExcelWriter('CS2 stats for stream.xlsx', engine='xlsxwriter') as writer:
                self.total_score_df.to_excel(writer, sheet_name='CS2', index=False, startrow=0, startcol=0)
                self.players_stats_df.to_excel(writer, sheet_name='CS2', index=False, startrow=len(self.total_score_df) + 2, startcol=0)
